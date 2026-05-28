import os
import io
import re
import datetime as dt
from pathlib import Path
from typing import Optional, List
from fastapi import (
    FastAPI,
    Depends,
    HTTPException,
    UploadFile,
    File,
    Form,
    Request,
    status,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import OAuth2PasswordBearer
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr
from sqlalchemy import text
from sqlalchemy.orm import Session
import bcrypt
from jose import jwt, JWTError
from openpyxl import load_workbook

# Engine, SessionLocal, and Base live in backend.database so they can be
# shared with Alembic (see alembic/env.py). Importing here triggers `.env`
# loading and DATABASE_URL validation as a side effect.
from backend.database import engine, SessionLocal, get_db  # noqa: F401
from backend import models  # noqa: F401  # registers ORM models on Base.metadata

SECRET_KEY = os.environ.get("SESSION_SECRET")
if not SECRET_KEY:
    raise RuntimeError("SESSION_SECRET environment variable is required")
ALGORITHM = "HS256"
TOKEN_EXPIRE_HOURS = 24 * 7
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode("utf-8")[:72], bcrypt.gensalt()).decode("utf-8")


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode("utf-8")[:72], h.encode("utf-8"))
    except Exception:
        return False


app = FastAPI(title="PROverify API")


# Schema is now managed by Alembic — run `alembic upgrade head` to apply
# pending migrations. See alembic/README.md for details.


app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r".*",
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# NOTE: `get_db` is imported from backend.database at the top of this file.


def seed_admin():
    """Idempotently ensure the bootstrap admin exists.

    Credentials are taken from `INITIAL_ADMIN_EMAIL` / `INITIAL_ADMIN_PASSWORD`
    so production deployments can set their own and never rely on a default.
    If both are unset, seeding is skipped — in that case the admins table must
    be populated out-of-band before anyone can log in.

    All other tables (brands, product_codes, upload_batches, verification_logs)
    are intentionally left empty so each environment starts clean and is
    populated only through the admin UI.
    """
    email = os.environ.get("INITIAL_ADMIN_EMAIL")
    password = os.environ.get("INITIAL_ADMIN_PASSWORD")
    if not email or not password:
        return
    db = SessionLocal()
    try:
        row = db.execute(
            text("SELECT id FROM admins WHERE email=:e"), {"e": email}
        ).first()
        if not row:
            db.execute(
                text("INSERT INTO admins (email, password_hash) VALUES (:e,:p)"),
                {"e": email, "p": hash_password(password)},
            )
            db.commit()
    finally:
        db.close()


seed_admin()


def create_token(email: str) -> str:
    payload = {
        "sub": email,
        "exp": dt.datetime.utcnow() + dt.timedelta(hours=TOKEN_EXPIRE_HOURS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def current_admin(
    token: Optional[str] = Depends(oauth2_scheme), db: Session = Depends(get_db)
):
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    row = db.execute(
        text("SELECT id, email FROM admins WHERE email=:e"), {"e": email}
    ).first()
    if not row:
        raise HTTPException(status_code=401, detail="Admin not found")
    return {"id": row[0], "email": row[1]}


# ---------- Schemas ----------
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class BrandIn(BaseModel):
    name: str
    primary_color: Optional[str] = "#1b5e20"
    desktop_image: Optional[str] = None
    mobile_image: Optional[str] = None
    is_active: Optional[bool] = True


# ---------- Auth ----------
@app.post("/api/auth/login")
def login(data: LoginIn, db: Session = Depends(get_db)):
    row = db.execute(
        text("SELECT id, email, password_hash FROM admins WHERE email=:e"),
        {"e": data.email},
    ).first()
    if not row or not verify_password(data.password, row[2]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_token(row[1])
    return {"access_token": token, "token_type": "bearer", "email": row[1]}


@app.get("/api/auth/me")
def me(admin=Depends(current_admin)):
    return admin


# ---------- Helpers ----------
def slugify(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", name.strip().lower()).strip("-")
    return s or "brand"


def unique_slug(db: Session, base: str, ignore_id: Optional[int] = None) -> str:
    slug = base
    i = 2
    while True:
        q = "SELECT id FROM brands WHERE slug=:s"
        params = {"s": slug}
        if ignore_id:
            q += " AND id<>:i"
            params["i"] = ignore_id
        if not db.execute(text(q), params).first():
            return slug
        slug = f"{base}-{i}"
        i += 1


# ---------- Brands ----------
@app.get("/api/brands")
def list_brands(
    active_only: bool = False,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    q = "SELECT id, name, slug, primary_color, desktop_image, mobile_image, is_active, created_at, updated_at FROM brands"
    if active_only:
        q += " WHERE is_active=TRUE"
    q += " ORDER BY id ASC"
    rows = db.execute(text(q)).all()
    return [
        {
            "id": r[0],
            "name": r[1],
            "slug": r[2],
            "primary_color": r[3],
            "desktop_image": r[4],
            "mobile_image": r[5],
            "is_active": r[6],
            "created_at": r[7].isoformat() if r[7] else None,
            "updated_at": r[8].isoformat() if r[8] else None,
        }
        for r in rows
    ]


@app.post("/api/brands")
def create_brand(
    data: BrandIn, db: Session = Depends(get_db), admin=Depends(current_admin)
):
    slug = unique_slug(db, slugify(data.name))
    row = db.execute(
        text("""INSERT INTO brands (name, slug, primary_color, desktop_image, mobile_image, is_active)
                VALUES (:n,:s,:c,:d,:m,:a) RETURNING id"""),
        {
            "n": data.name,
            "s": slug,
            "c": data.primary_color,
            "d": data.desktop_image,
            "m": data.mobile_image,
            "a": data.is_active,
        },
    ).first()
    db.commit()
    return {"id": row[0], "slug": slug}


@app.put("/api/brands/{brand_id}")
def update_brand(
    brand_id: int,
    data: BrandIn,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    existing = db.execute(
        text("SELECT id FROM brands WHERE id=:i"), {"i": brand_id}
    ).first()
    if not existing:
        raise HTTPException(404, "Brand not found")
    slug = unique_slug(db, slugify(data.name), ignore_id=brand_id)
    db.execute(
        text("""UPDATE brands SET name=:n, slug=:s, primary_color=:c, desktop_image=:d,
                mobile_image=:m, is_active=:a, updated_at=NOW() WHERE id=:i"""),
        {
            "n": data.name,
            "s": slug,
            "c": data.primary_color,
            "d": data.desktop_image,
            "m": data.mobile_image,
            "a": data.is_active,
            "i": brand_id,
        },
    )
    db.commit()
    return {"ok": True}


@app.patch("/api/brands/{brand_id}/toggle")
def toggle_brand(
    brand_id: int, db: Session = Depends(get_db), admin=Depends(current_admin)
):
    db.execute(
        text(
            "UPDATE brands SET is_active = NOT is_active, updated_at=NOW() WHERE id=:i"
        ),
        {"i": brand_id},
    )
    db.commit()
    return {"ok": True}


@app.delete("/api/brands/{brand_id}")
def delete_brand(
    brand_id: int, db: Session = Depends(get_db), admin=Depends(current_admin)
):
    db.execute(text("DELETE FROM brands WHERE id=:i"), {"i": brand_id})
    db.commit()
    return {"ok": True}


# ---------- Image Upload (saved to disk, returns public path) ----------
_UPLOADS_DIR = Path(__file__).resolve().parent.parent / "uploads"
_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
(_UPLOADS_DIR / "brands").mkdir(parents=True, exist_ok=True)

_ALLOWED_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
_MAX_IMAGE_BYTES = 10 * 1024 * 1024  # 10 MB


def _sniff_image_kind(data: bytes) -> Optional[str]:
    """Return 'png'|'jpg'|'gif'|'webp' from the file's magic bytes, or None."""
    if len(data) < 12:
        return None
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if data[:3] == b"\xff\xd8\xff":
        return "jpg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "webp"
    return None


@app.post("/api/uploads/image")
async def upload_image(
    file: UploadFile = File(...),
    folder: str = Form("brands"),
    admin=Depends(current_admin),
):
    safe_folder = re.sub(r"[^a-z0-9_-]", "", folder.lower()) or "misc"
    ext = Path(file.filename or "").suffix.lower()
    if ext not in _ALLOWED_IMAGE_EXT:
        raise HTTPException(400, f"Unsupported image type '{ext}'. Allowed: {', '.join(sorted(_ALLOWED_IMAGE_EXT))}")
    data = await file.read()
    if len(data) > _MAX_IMAGE_BYTES:
        raise HTTPException(413, f"Image too large ({len(data)/1024/1024:.1f} MB). Max 10 MB.")
    kind = _sniff_image_kind(data)
    if not kind:
        raise HTTPException(400, "File is not a recognized image (PNG, JPG, GIF, WEBP).")
    # Re-derive extension from real type so spoofed extensions can't change it.
    canonical_ext = ".jpg" if kind == "jpg" else f".{kind}"
    ext = canonical_ext
    import secrets
    name = f"{dt.datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(8)}{ext}"
    target_dir = _UPLOADS_DIR / safe_folder
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / name
    target.write_bytes(data)
    return {"url": f"/uploads/{safe_folder}/{name}", "size": len(data)}


# ---------- Codes Upload (streaming + batched insert) ----------
@app.post("/api/codes/upload")
async def upload_codes(
    brand_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    brand = db.execute(
        text("SELECT id, slug FROM brands WHERE id=:i AND is_active=TRUE"),
        {"i": brand_id},
    ).first()
    if not brand:
        raise HTTPException(400, "Brand not found or inactive")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")

    # Stream upload to a temp file so we don't hold the whole xlsx in memory.
    import tempfile, shutil, os as _os

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    batch_id = None
    try:
        try:
            shutil.copyfileobj(file.file, tmp, length=1024 * 1024)
        finally:
            tmp.close()

        try:
            wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Could not read xlsx: {e}")
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise HTTPException(400, "Empty file")
        code_idx = None
        for i, h in enumerate(header_row):
            if h and str(h).strip().lower() == "code":
                code_idx = i
                break
        if code_idx is None:
            raise HTTPException(400, 'Excel must have a "Code" column header')

        # Create batch row
        today = dt.datetime.utcnow().strftime("%Y%m%d")
        count_today = db.execute(
            text(
                "SELECT COUNT(*) FROM upload_batches WHERE brand_id=:b AND DATE(created_at)=CURRENT_DATE"
            ),
            {"b": brand_id},
        ).scalar()
        batch_number = f"{brand[1].upper()}-{today}-{int(count_today) + 1:04d}"
        batch_row = db.execute(
            text("""INSERT INTO upload_batches (brand_id, batch_number, file_name, codes_uploaded)
                    VALUES (:b,:n,:f,0) RETURNING id"""),
            {"b": brand_id, "n": batch_number, "f": file.filename},
        ).first()
        batch_id = batch_row[0]
        db.commit()

        # Streaming batched insert with duplicate tracking.
        # In-file dupes (same code appears twice in the xlsx) and DB dupes
        # (code already exists for this brand from a prior upload) are skipped
        # via ON CONFLICT, and counted separately.
        from psycopg2.extras import execute_values

        CHUNK = 5000
        SAMPLE_LIMIT = 15
        chunk: List[tuple] = []
        seen_in_file: set = set()
        total_rows = 0
        in_file_dupes = 0
        db_dupes = 0
        file_dup_samples: List[str] = []
        db_dup_samples: List[str] = []
        inserted = 0

        def flush(cur, chunk_):
            nonlocal inserted, db_dupes
            if not chunk_:
                return
            result = execute_values(
                cur,
                "INSERT INTO product_codes (code, brand_id, batch_id) VALUES %s "
                "ON CONFLICT (brand_id, code) DO NOTHING RETURNING code",
                chunk_,
                template="(%s,%s,%s)",
                fetch=True,
            )
            inserted_codes = {r[0] for r in result}
            inserted += len(inserted_codes)
            for c, _, _ in chunk_:
                if c not in inserted_codes:
                    db_dupes += 1
                    if len(db_dup_samples) < SAMPLE_LIMIT:
                        db_dup_samples.append(c)

        conn = engine.raw_connection()
        try:
            try:
                cur = conn.cursor()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or code_idx >= len(row):
                        continue
                    val = row[code_idx]
                    if val is None:
                        continue
                    code = str(val).strip()
                    if not code:
                        continue
                    total_rows += 1
                    if code in seen_in_file:
                        in_file_dupes += 1
                        if len(file_dup_samples) < SAMPLE_LIMIT:
                            file_dup_samples.append(code)
                        continue
                    seen_in_file.add(code)
                    chunk.append((code, brand_id, batch_id))
                    if len(chunk) >= CHUNK:
                        flush(cur, chunk)
                        chunk = []
                flush(cur, chunk)
                conn.commit()
                cur.close()
            except Exception:
                conn.rollback()
                raise
        finally:
            conn.close()

        db.execute(
            text("UPDATE upload_batches SET codes_uploaded=:c WHERE id=:i"),
            {"c": inserted, "i": batch_id},
        )
        db.commit()
        return {
            "batch_id": batch_id,
            "batch_number": batch_number,
            "codes_uploaded": inserted,
            "rows_processed": total_rows,
            "duplicates_in_file": in_file_dupes,
            "duplicates_in_db": db_dupes,
            "duplicate_samples_in_file": file_dup_samples,
            "duplicate_samples_in_db": db_dup_samples,
        }

    except HTTPException:
        if batch_id is not None:
            try:
                db.execute(
                    text("DELETE FROM upload_batches WHERE id=:i"), {"i": batch_id}
                )
                db.commit()
            except Exception:
                db.rollback()
        raise
    except Exception as e:
        if batch_id is not None:
            try:
                db.execute(
                    text("DELETE FROM upload_batches WHERE id=:i"), {"i": batch_id}
                )
                db.commit()
            except Exception:
                db.rollback()
        raise HTTPException(500, f"Upload failed: {e}")
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.get("/api/codes/sample")
def codes_sample():
    """Return a tiny sample .xlsx with a Code column."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Codes"
    ws.append(["Code"])
    for i in range(1, 11):
        ws.append([f"SAMPLE-{i:04d}"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="sample_codes.xlsx"'},
    )


# ---------- Batches & Codes listing ----------
@app.get("/api/batches")
def list_batches(
    search: Optional[str] = None,
    brand_id: Optional[int] = None,
    limit: int = 25,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    limit = min(max(limit, 1), 200)
    offset = max(offset, 0)
    where = ["1=1"]
    params: dict = {}
    if brand_id:
        where.append("b.brand_id = :bid")
        params["bid"] = brand_id
    if search:
        where.append(
            "(b.batch_number ILIKE :q OR b.file_name ILIKE :q OR br.name ILIKE :q)"
        )
        params["q"] = f"%{search}%"
    w = " AND ".join(where)
    total = (
        db.execute(
            text(f"""
        SELECT COUNT(*) FROM upload_batches b JOIN brands br ON br.id=b.brand_id WHERE {w}
    """),
            params,
        ).scalar()
        or 0
    )
    params["lim"] = limit
    params["off"] = offset
    rows = db.execute(
        text(f"""
        SELECT b.id, b.batch_number, b.file_name, b.codes_uploaded, b.created_at,
               br.id, br.name, br.slug
        FROM upload_batches b
        JOIN brands br ON br.id=b.brand_id
        WHERE {w}
        ORDER BY b.created_at DESC
        LIMIT :lim OFFSET :off
    """),
        params,
    ).all()
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r[0],
                "batch_number": r[1],
                "file_name": r[2],
                "codes_uploaded": r[3],
                "created_at": r[4].isoformat() if r[4] else None,
                "brand_id": r[5],
                "brand_name": r[6],
                "brand_slug": r[7],
            }
            for r in rows
        ],
    }


@app.get("/api/batches/{batch_id}")
def batch_detail(
    batch_id: int, db: Session = Depends(get_db), admin=Depends(current_admin)
):
    row = db.execute(
        text("""
        SELECT b.id, b.batch_number, b.file_name, b.codes_uploaded, b.created_at,
               br.id, br.name, br.slug
        FROM upload_batches b JOIN brands br ON br.id=b.brand_id
        WHERE b.id=:i
    """),
        {"i": batch_id},
    ).first()
    if not row:
        raise HTTPException(404, "Batch not found")
    # Aggregate verification stats: count valid-scan logs against codes in this batch.
    stats = db.execute(
        text("""
        SELECT
            COUNT(v.id) FILTER (WHERE v.is_valid = TRUE) AS total_verifications,
            COUNT(DISTINCT v.code) FILTER (WHERE v.is_valid = TRUE) AS codes_verified,
            MAX(v.created_at) FILTER (WHERE v.is_valid = TRUE) AS last_verified_at
        FROM product_codes pc
        LEFT JOIN verification_logs v
          ON v.code = pc.code AND v.brand_id = pc.brand_id
        WHERE pc.batch_id = :i
    """),
        {"i": batch_id},
    ).first()
    return {
        "id": row[0],
        "batch_number": row[1],
        "file_name": row[2],
        "codes_uploaded": row[3],
        "created_at": row[4].isoformat() if row[4] else None,
        "brand_id": row[5],
        "brand_name": row[6],
        "brand_slug": row[7],
        "total_verifications": int(stats[0] or 0),
        "codes_verified": int(stats[1] or 0),
        "last_verified_at": stats[2].isoformat() if stats[2] else None,
    }


@app.get("/api/batches/{batch_id}/codes")
def batch_codes(
    batch_id: int,
    search: Optional[str] = None,
    verified_only: bool = False,
    limit: int = 25,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    limit = min(max(limit, 1), 200)
    offset = max(offset, 0)
    exists = db.execute(
        text("SELECT 1 FROM upload_batches WHERE id=:i"), {"i": batch_id}
    ).first()
    if not exists:
        raise HTTPException(404, "Batch not found")
    where = ["pc.batch_id = :i"]
    params: dict = {"i": batch_id}
    if search:
        where.append("pc.code ILIKE :q")
        params["q"] = f"%{search}%"
    having = ""
    if verified_only:
        having = "HAVING COUNT(v.id) FILTER (WHERE v.is_valid = TRUE) > 0"
    w = " AND ".join(where)
    total = (
        db.execute(
            text(f"""
        SELECT COUNT(*) FROM (
            SELECT pc.id
            FROM product_codes pc
            LEFT JOIN verification_logs v ON v.code = pc.code AND v.brand_id = pc.brand_id
            WHERE {w}
            GROUP BY pc.id
            {having}
        ) t
    """),
            params,
        ).scalar()
        or 0
    )
    params["lim"] = limit
    params["off"] = offset
    rows = db.execute(
        text(f"""
        SELECT pc.id, pc.code, pc.created_at,
               COUNT(v.id) FILTER (WHERE v.is_valid = TRUE) AS verification_count,
               MAX(v.created_at) FILTER (WHERE v.is_valid = TRUE) AS last_verified_at
        FROM product_codes pc
        LEFT JOIN verification_logs v ON v.code = pc.code AND v.brand_id = pc.brand_id
        WHERE {w}
        GROUP BY pc.id
        {having}
        ORDER BY verification_count DESC, pc.id ASC
        LIMIT :lim OFFSET :off
    """),
        params,
    ).all()
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r[0],
                "code": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "verification_count": int(r[3] or 0),
                "last_verified_at": r[4].isoformat() if r[4] else None,
            }
            for r in rows
        ],
    }


@app.get("/api/codes/{code_id}/logs")
def code_logs(
    code_id: int,
    search: Optional[str] = None,
    valid: Optional[str] = None,
    limit: int = 25,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    limit = min(max(limit, 1), 500)
    offset = max(offset, 0)
    pc = db.execute(
        text("""
        SELECT pc.id, pc.code, pc.brand_id, br.name, br.slug
        FROM product_codes pc LEFT JOIN brands br ON br.id = pc.brand_id
        WHERE pc.id = :i
    """),
        {"i": code_id},
    ).first()
    if not pc:
        raise HTTPException(404, "Code not found")
    where = ["code = :c", "brand_id = :b"]
    params: dict = {"c": pc[1], "b": pc[2]}
    if search:
        where.append("(CAST(ip_address AS TEXT) ILIKE :q OR user_agent ILIKE :q)")
        params["q"] = f"%{search}%"
    if valid == "true":
        where.append("is_valid = TRUE")
    elif valid == "false":
        where.append("is_valid = FALSE")
    w = " AND ".join(where)
    total = (
        db.execute(
            text(f"SELECT COUNT(*) FROM verification_logs WHERE {w}"), params
        ).scalar()
        or 0
    )
    params["l"] = limit
    params["o"] = offset
    rows = db.execute(
        text(f"""
        SELECT id, is_valid, created_at, ip_address, user_agent
        FROM verification_logs
        WHERE {w}
        ORDER BY created_at DESC, id DESC
        LIMIT :l OFFSET :o
    """),
        params,
    ).all()
    return {
        "code": pc[1],
        "brand_name": pc[3],
        "brand_slug": pc[4],
        "total": int(total),
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r[0],
                "is_valid": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "ip": r[3],
                "user_agent": r[4],
            }
            for r in rows
        ],
    }


@app.delete("/api/batches/{batch_id}")
def delete_batch(
    batch_id: int, db: Session = Depends(get_db), admin=Depends(current_admin)
):
    db.execute(text("DELETE FROM upload_batches WHERE id=:i"), {"i": batch_id})
    db.commit()
    return {"ok": True}


@app.get("/api/codes")
def list_codes(
    brand_id: Optional[int] = None,
    batch_id: Optional[int] = None,
    search: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    limit = min(max(limit, 1), 500)
    where = ["1=1"]
    params = {}
    if brand_id:
        where.append("pc.brand_id=:b")
        params["b"] = brand_id
    if batch_id:
        where.append("pc.batch_id=:bt")
        params["bt"] = batch_id
    if search:
        where.append("pc.code ILIKE :s")
        params["s"] = f"%{search}%"
    w = " AND ".join(where)
    total = db.execute(
        text(f"SELECT COUNT(*) FROM product_codes pc WHERE {w}"), params
    ).scalar()
    params["lim"] = limit
    params["off"] = offset
    rows = db.execute(
        text(f"""
        SELECT pc.id, pc.code, pc.created_at, br.name
        FROM product_codes pc JOIN brands br ON br.id=pc.brand_id
        WHERE {w}
        ORDER BY pc.id DESC LIMIT :lim OFFSET :off
    """),
        params,
    ).all()
    return {
        "total": int(total or 0),
        "items": [
            {
                "id": r[0],
                "code": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "brand_name": r[3],
            }
            for r in rows
        ],
    }


# ---------- Dashboard ----------
@app.get("/api/dashboard/stats")
def dashboard_stats(db: Session = Depends(get_db), admin=Depends(current_admin)):
    total_codes = db.execute(text("SELECT COUNT(*) FROM product_codes")).scalar() or 0
    total_verifs = (
        db.execute(text("SELECT COUNT(*) FROM verification_logs")).scalar() or 0
    )
    total_batches = (
        db.execute(text("SELECT COUNT(*) FROM upload_batches")).scalar() or 0
    )
    total_brands = db.execute(text("SELECT COUNT(*) FROM brands")).scalar() or 0

    codes_per_brand = db.execute(
        text("""
        SELECT br.id, br.name, COUNT(pc.id)
        FROM brands br LEFT JOIN product_codes pc ON pc.brand_id=br.id
        GROUP BY br.id, br.name ORDER BY br.id
    """)
    ).all()
    verifs_per_brand = db.execute(
        text("""
        SELECT br.id, br.name, COUNT(v.id)
        FROM brands br LEFT JOIN verification_logs v ON v.brand_id=br.id
        GROUP BY br.id, br.name ORDER BY br.id
    """)
    ).all()
    return {
        "totals": {
            "codes": int(total_codes),
            "verifications": int(total_verifs),
            "batches": int(total_batches),
            "brands": int(total_brands),
        },
        "codes_per_brand": [
            {"brand": r[1], "count": int(r[2])} for r in codes_per_brand
        ],
        "verifications_per_brand": [
            {"brand": r[1], "count": int(r[2])} for r in verifs_per_brand
        ],
    }


@app.get("/api/public/stats")
def public_stats(db: Session = Depends(get_db)):
    codes = db.execute(text("SELECT COUNT(*) FROM product_codes")).scalar() or 0
    verifs = (
        db.execute(
            text("SELECT COUNT(*) FROM verification_logs WHERE is_valid=TRUE")
        ).scalar()
        or 0
    )
    return {"codes": int(codes), "verifications": int(verifs), "uptime": "99.9%"}


@app.get("/api/public/brands/{slug}")
def public_brand(slug: str, db: Session = Depends(get_db)):
    r = db.execute(
        text(
            "SELECT id, name, slug, primary_color, desktop_image, mobile_image FROM brands WHERE slug=:s AND is_active=TRUE"
        ),
        {"s": slug},
    ).first()
    if not r:
        raise HTTPException(404, "Brand not found")
    return {
        "id": r[0],
        "name": r[1],
        "slug": r[2],
        "primary_color": r[3],
        "desktop_image": r[4],
        "mobile_image": r[5],
    }


class VerifyIn(BaseModel):
    slug: str
    code: str


_rate_buckets: dict = {}


def _rate_limit(ip: Optional[str], max_per_minute: int = 30) -> bool:
    if not ip:
        return True
    now = dt.datetime.utcnow().timestamp()
    window = 60.0
    hits = [t for t in _rate_buckets.get(ip, []) if now - t < window]
    if len(hits) >= max_per_minute:
        _rate_buckets[ip] = hits
        return False
    hits.append(now)
    _rate_buckets[ip] = hits
    return True


@app.post("/api/public/verify")
def verify_code(data: VerifyIn, request: Request, db: Session = Depends(get_db)):
    code = (data.code or "").strip()
    slug = (data.slug or "").strip()
    if not code or len(code) > 255:
        raise HTTPException(400, "Invalid code")
    if not slug or len(slug) > 255:
        raise HTTPException(400, "Invalid slug")
    ip = request.client.host if request.client else None
    if not _rate_limit(ip):
        raise HTTPException(429, "Too many requests, please slow down")
    ua = request.headers.get("user-agent")

    brand = db.execute(
        text("SELECT id, name FROM brands WHERE slug=:s AND is_active=TRUE"),
        {"s": slug},
    ).first()
    if not brand:
        raise HTTPException(404, "Brand not found")

    # Lock the product_code row so concurrent verifies for the same code
    # serialize, preventing two requests from both being classified as "first".
    pc = db.execute(
        text("SELECT id FROM product_codes WHERE code=:c AND brand_id=:b FOR UPDATE"),
        {"c": code, "b": brand[0]},
    ).first()
    if not pc:
        db.execute(
            text("""INSERT INTO verification_logs (code, brand_id, ip_address, user_agent, is_valid)
                    VALUES (:c,:b,:ip,:ua,FALSE)"""),
            {"c": code, "b": brand[0], "ip": ip, "ua": ua},
        )
        db.commit()
        return {"status": "invalid", "brand": brand[1]}

    prior = db.execute(
        text("""SELECT created_at FROM verification_logs
                WHERE code=:c AND brand_id=:b AND is_valid=TRUE
                ORDER BY created_at ASC"""),
        {"c": code, "b": brand[0]},
    ).all()
    now = dt.datetime.utcnow()
    db.execute(
        text("""INSERT INTO verification_logs (code, brand_id, ip_address, user_agent, is_valid)
                VALUES (:c,:b,:ip,:ua,TRUE)"""),
        {"c": code, "b": brand[0], "ip": ip, "ua": ua},
    )
    db.commit()
    if not prior:
        return {
            "status": "first",
            "brand": brand[1],
            "code": code,
            "verified_at": now.isoformat(),
            "history": [],
        }
    return {
        "status": "repeat",
        "brand": brand[1],
        "code": code,
        "first_verified_at": prior[0][0].isoformat(),
        "current_scan_at": now.isoformat(),
        "history": [r[0].isoformat() for r in prior],
    }


@app.get("/api/activity")
def activity(
    brand_id: Optional[int] = None,
    valid: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    where = ["1=1"]
    params: dict = {"l": max(1, min(limit, 1000))}
    if brand_id:
        where.append("v.brand_id=:b")
        params["b"] = brand_id
    if valid == "true":
        where.append("v.is_valid=TRUE")
    elif valid == "false":
        where.append("v.is_valid=FALSE")
    rows = db.execute(
        text(f"""SELECT v.id, v.code, v.is_valid, v.created_at, v.ip_address, br.name
                 FROM verification_logs v LEFT JOIN brands br ON br.id=v.brand_id
                 WHERE {" AND ".join(where)}
                 ORDER BY v.id DESC LIMIT :l"""),
        params,
    ).all()
    return [
        {
            "id": r[0],
            "code": r[1],
            "is_valid": r[2],
            "created_at": r[3].isoformat() if r[3] else None,
            "ip": r[4],
            "brand": r[5],
        }
        for r in rows
    ]


@app.post("/api/admin/reset-data")
def reset_data(
    confirm: str,
    db: Session = Depends(get_db),
    admin=Depends(current_admin),
):
    """Destructive: wipe ALL brands, codes, batches and verification logs.

    Keeps the `admins` table intact. Requires the literal confirmation string
    ``RESET-ALL-DATA`` to be sent as a query parameter to prevent accidents.
    """
    if confirm != "RESET-ALL-DATA":
        raise HTTPException(400, "Missing or invalid confirmation token")
    db.execute(
        text(
            "TRUNCATE TABLE verification_logs, product_codes, upload_batches, brands "
            "RESTART IDENTITY CASCADE"
        )
    )
    db.commit()
    counts = {}
    for tbl in (
        "admins",
        "brands",
        "product_codes",
        "upload_batches",
        "verification_logs",
    ):
        counts[tbl] = db.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
    return {"ok": True, "wiped_by": admin["email"], "row_counts": counts}


@app.get("/api")
def api_root():
    return {"service": "PROverify API", "status": "ok"}


# Serve the built React SPA. In production the deployment build step runs
# `npm run build` which writes static assets into `frontend/dist`. When that
# directory exists we mount it and add a catch-all so client-side routes
# (/login, /brands, /verify/<slug>, ...) all return index.html.
app.mount("/uploads", StaticFiles(directory=_UPLOADS_DIR), name="uploads")

_FRONTEND_DIST = Path(__file__).resolve().parent.parent / "frontend" / "dist"
if _FRONTEND_DIST.is_dir():
    _INDEX_HTML = _FRONTEND_DIST / "index.html"
    app.mount(
        "/assets",
        StaticFiles(directory=_FRONTEND_DIST / "assets"),
        name="assets",
    )

    @app.get("/")
    def _spa_root():
        return FileResponse(_INDEX_HTML)

    @app.get("/{full_path:path}")
    def _spa_catch_all(full_path: str):
        # Never hijack API or uploads routes — let FastAPI return its own 404 for them.
        if full_path.startswith("api/") or full_path.startswith("uploads/"):
            raise HTTPException(status_code=404, detail="Not Found")
        # Serve a real file from dist if one exists (favicon.svg, vite.svg, ...).
        candidate = _FRONTEND_DIST / full_path
        if candidate.is_file():
            return FileResponse(candidate)
        # Otherwise fall back to the SPA shell so React Router can handle it.
        return FileResponse(_INDEX_HTML)
else:

    @app.get("/")
    def _root_dev():
        return {
            "service": "PROverify API",
            "status": "ok",
            "note": "frontend/dist not built; run `npm run build` for production",
        }
